import sys

import openpyxl

#workbook = openpyxl.Workbook()

name = 'Yogesh'
email = 'yogymax@gmail.com'
age = 33

statement = "My name is Yogesh and email address is yogymax@gmail.com, age 33"
statement1 = "My name is {} and email address is {}, age {}".format(name,email,age)  # string ko format #position
statement2 = "My name is {2} and email address is {0}, age {1}".format(email,age,name) # index
statement3 = "My name is {x} and email address is {z}, age {y}".format(x=name,y=age,z=email) # name change hua
statement4 = f"My name is {name} and email address is {email}, age {age}"

print(statement1)
print(statement2)
print(statement3)
print(statement4)

#x = "name"
#age = 10

#x+age



class Book:

    def __init__(self, bkid,bknm,bkprc,bkauth,bkqty):
        self.book_id = bkid
        self.book_name = bknm
        self.book_price = bkprc
        self.book_author = bkauth
        self.book_quantity = bkqty

    def __str__(self):               #callback       # str ==?  str == which represents to the object..
        return f'''{self.__dict__}'''

    def __repr__(self):  # repr--str-- complex of complex  # call back
        return str(self)

    def display_book_information(self):
        print('inside display information')

#callback methods ==?

id = int(input('Enter Book Id : '))
name = input('Enter Book Name : ')
qty = int(input('Enter Book Qty : '))
price = float(input('Enter Book Price  :'))
author = input('Enter Author Name : ')
book = Book(id,name,price,author,qty)
print(book) # user is going to enter the values -

import openpyxl


def read_data_from_excel():
    workbook = openpyxl.load_workbook('book.xlsx')
    sheet = workbook['book_information']
    print(sheet['A2'].value)
    print(sheet['D2'].value)


read_data_from_excel()

print('=================================')
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('book_information')
sheet['A1'] = 'BOOK_ID'
sheet['B1'] = 'BOOK_NAME'
sheet['C1'] = 'BOOK_PRICE'
sheet['D1'] = 'BOOK_AUTHOR'
sheet['E1'] = 'BOOK_QTY'
sheet['A2'] = book.book_id
sheet['B2'] = book.book_name
sheet['C2'] = book.book_price
sheet['D2'] = book.book_author
sheet['E2'] = book.book_quantity
workbook.save('book.xlsx')
print('Book Information recorded into excel file...')




sys.exit(0)
b1 = Book(101,'Python',2345.5,'Mr XXXX',23)
b2 = Book(102,'Python-Web',4234.5,'Mr XXXX1',63)
b3 = Book(103,'Python-DS',3234.5,'Mr XXXX2',28)

print(b1) # b1 -- is the ref -- to the object of type -- Book -- who defined this book -- developer
print(id(b1))  #2368758609808 === id method --. hexadecimal to decimal-->

print(b2)

b1.display_book_information()   #
b2.display_book_information()

print(b3)

book_list = [b1,b2,b3]  # complex of complex..
print(book_list)